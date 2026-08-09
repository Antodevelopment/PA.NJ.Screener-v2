"""Build the app's screened parcel dataset from official NJ and Chester County records."""

from __future__ import annotations

import json
import io
import re
import shutil
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUTPUT = ROOT / "app" / "data" / "real-parcels.json"
GEOCODE_CACHE = DATA / "nj-geocodes.json"
CHESTER_GEOCODE_CACHE = DATA / "chester-geocodes.json"
CHECKED = datetime.now(timezone.utc).date().isoformat()
CHESTER_PARCELS_URL = "https://services.arcgis.com/G4S1dGvn7PIgYd6Y/arcgis/rest/services/Parcels_owners/FeatureServer/0/query"
NJ_CODES_URL = "https://www.nj.gov/treasury/taxation/documents/excel/lpt/CDcodes.xlsx"
NJ_ARCHIVE_URL = "https://www.nj.gov/treasury/taxation/pdf/lpt/modiv-2026.zip"
NJ_ARCHIVE = DATA / "modiv-2026.zip"
NJ_ARCHIVE_META = DATA / "modiv-2026.meta.json"

CHESTER_MUNICIPALITIES = {
    1: "West Chester", 2: "Malvern", 3: "Kennett Square", 4: "Avondale",
    5: "West Grove", 6: "Oxford", 7: "Atglen", 8: "Parkesburg",
    9: "South Coatesville", 10: "Modena", 11: "Downingtown", 12: "Honey Brook Borough",
    13: "Elverson", 14: "Spring City", 15: "Phoenixville", 16: "Coatesville",
    17: "North Coventry", 18: "East Coventry", 19: "Warwick", 20: "South Coventry",
    21: "East Vincent", 22: "Honey Brook Township", 23: "West Nantmeal", 24: "East Nantmeal",
    25: "West Vincent", 26: "East Pikeland", 27: "Schuylkill", 28: "West Caln",
    29: "West Brandywine", 30: "East Brandywine", 31: "Wallace", 32: "Upper Uwchlan",
    33: "Uwchlan", 34: "West Pikeland", 35: "Charlestown", 36: "West Sadsbury",
    37: "Sadsbury", 38: "Valley", 39: "Caln", 40: "East Caln",
    41: "West Whiteland", 42: "East Whiteland", 43: "Tredyffrin", 44: "West Fallowfield",
    45: "Highland", 46: "Londonderry", 47: "East Fallowfield", 48: "West Marlborough",
    49: "Newlin", 50: "West Bradford", 51: "East Bradford", 52: "West Goshen",
    53: "East Goshen", 54: "Willistown", 55: "Easttown", 56: "Lower Oxford",
    57: "Upper Oxford", 58: "Penn", 59: "London Grove", 60: "New Garden",
    61: "East Marlborough", 62: "Kennett", 63: "Pocopson", 64: "Pennsbury",
    65: "Birmingham", 66: "Thornbury", 67: "Westtown", 68: "West Nottingham",
    69: "East Nottingham", 70: "Elk", 71: "New London", 72: "Franklin",
    73: "London Britain",
}

CHESTER_URBAN_CENTERS = {
    "Atglen", "Avondale", "Coatesville", "Downingtown", "Elverson", "Honey Brook Borough",
    "Kennett Square", "Modena", "Oxford", "Phoenixville", "South Coatesville", "Spring City",
    "West Chester", "West Grove",
}

LAND_USE_NAMES = {
    "V-10": "Vacant residential land", "V-11": "Vacant commercial land",
    "C-70": "Commercial garage / shop", "C-80": "Warehouse",
    "C-96": "Commercial outbuilding only", "M-20": "Light industrial",
}


def fw(line: str, start: int, end: int) -> str:
    return line[start - 1 : end].strip()


def number(value: str) -> int:
    try:
        return int(float(value.strip() or 0))
    except ValueError:
        return 0


def deed_year(raw: str) -> int:
    if len(raw) != 6 or raw == "000000":
        return 0
    yy = int(raw[-2:])
    return 2000 + yy if yy <= 26 else 1900 + yy


def owner_type(name: str) -> str:
    n = name.upper()
    if "TRUST" in n or "ESTATE" in n or re.search(r"\bEST\b", n):
        return "Estate / trust"
    if any(x in n for x in (" LLC", " LP", " L P", " INC", " CORP", " ASSOC")):
        return "Business entity"
    if any(x in n for x in ("UNIVERSITY", "CITY OF", "AUTHORITY")):
        return "Institutional"
    return "Individual"


def load_nj_municipalities() -> dict[str, str]:
    with urllib.request.urlopen(NJ_CODES_URL, timeout=60) as response:
        workbook = openpyxl.load_workbook(io.BytesIO(response.read()), read_only=True, data_only=True)
    municipalities: dict[str, str] = {}
    for code, raw_name, _county in workbook[workbook.sheetnames[0]].iter_rows(min_row=2, values_only=True):
        if not code or not raw_name:
            continue
        name = str(raw_name).strip().title().replace(" Twp", " Township").replace(" Boro", " Borough")
        name = re.sub(r"\b(City|Township|Borough) \1\b", r"\1", name)
        municipalities[str(code).zfill(4)] = name
    if len(municipalities) < 500:
        raise RuntimeError(f"Official NJ code workbook returned only {len(municipalities)} municipalities")
    return municipalities


def refresh_nj_archive() -> Path:
    """Download the official NJ archive only when the state file has changed."""
    DATA.mkdir(parents=True, exist_ok=True)
    metadata = json.loads(NJ_ARCHIVE_META.read_text(encoding="utf-8")) if NJ_ARCHIVE_META.exists() else {}
    headers = {"User-Agent": "Matt's Deal Screener daily public-data refresh"}
    if NJ_ARCHIVE.exists() and metadata.get("etag"):
        headers["If-None-Match"] = metadata["etag"]
    if NJ_ARCHIVE.exists() and metadata.get("last_modified"):
        headers["If-Modified-Since"] = metadata["last_modified"]
    request = urllib.request.Request(NJ_ARCHIVE_URL, headers=headers)
    temporary = NJ_ARCHIVE.with_suffix(".download")
    try:
        with urllib.request.urlopen(request, timeout=120) as response, temporary.open("wb") as target:
            shutil.copyfileobj(response, target)
            updated_metadata = {
                "etag": response.headers.get("ETag", ""),
                "last_modified": response.headers.get("Last-Modified", ""),
                "downloaded": CHECKED,
            }
        with zipfile.ZipFile(temporary) as archive:
            if archive.testzip() is not None:
                raise RuntimeError("The downloaded NJ assessment archive is damaged")
        temporary.replace(NJ_ARCHIVE)
        NJ_ARCHIVE_META.write_text(json.dumps(updated_metadata, indent=2), encoding="utf-8")
        print("Downloaded the current official NJ MOD-IV archive")
    except urllib.error.HTTPError as error:
        if error.code != 304 or not NJ_ARCHIVE.exists():
            raise
        print("Official NJ MOD-IV archive is unchanged")
    except Exception:
        if not NJ_ARCHIVE.exists():
            raise
        print("NJ source was temporarily unavailable; using the last verified archive")
    finally:
        temporary.unlink(missing_ok=True)
    return NJ_ARCHIVE


def nj_candidates(stream, county: str, municipalities: dict[str, str]) -> list[dict]:
    rows = []
    for line in stream:
        district = fw(line, 1, 4)
        municipality = municipalities.get(district)
        prop_class = fw(line, 56, 58)
        location = fw(line, 59, 83)
        land = number(fw(line, 421, 429))
        improvement = number(fw(line, 430, 438))
        total = number(fw(line, 439, 447))
        if not municipality or not location or not 75_000 <= total <= 750_000:
            continue
        if prop_class not in {"1", "4A", "4B", "4C"}:
            continue
        ratio = improvement / land if land else 99
        if prop_class != "1" and ratio >= 0.45:
            continue
        block, lot, qualifier = fw(line, 5, 13), fw(line, 14, 22), fw(line, 23, 33)
        acres = number(fw(line, 119, 127)) / 10_000
        zoning = fw(line, 168, 171) or "Not supplied"
        mailing = ", ".join(x for x in (fw(line, 211, 235), fw(line, 236, 260), fw(line, 261, 269)) if x)
        year = number(fw(line, 416, 419))
        sale_year = deed_year(fw(line, 307, 312))
        delinquent = fw(line, 359, 359).upper() in {"Y", "D"}
        under = 95 if prop_class == "1" else max(60, min(90, round(92 - ratio * 50)))
        regulatory = 60 if zoning != "Not supplied" else 45
        motivation = 45 + (15 if sale_year and sale_year <= 2005 else 0) + (25 if delinquent else 0)
        strategic = 88 if total <= 300_000 else 82 if total <= 500_000 else 74
        parcel_id = f"{district}-{block}-{lot}" + (f"-{qualifier}" if qualifier else "")
        lot_sf = round(acres * 43_560) if acres else 0
        rows.append({
            "id": parcel_id,
            "owner": "Owner name redacted in NJ public export",
            "ownerType": "Public record redacted",
            "addressBase": location,
            "municipality": municipality,
            "mailing": mailing or "Not supplied",
            "lot": lot_sf,
            "building": 0,
            "year": year,
            "land": land,
            "improvement": improvement,
            "zoning": zoning,
            "flags": ["Official 2026 assessment", "Vacant land" if prop_class == "1" else f"Class {prop_class}"],
            "delinquency": "Delinquency flag present" if delinquent else "Not indicated in MOD-IV",
            "approvals": "Not researched",
            "factors": {"underutilization": under, "regulatory": regulatory, "motivation": min(100, motivation), "fit": strategic, "intelligence": 50},
            "reason": f"Official MOD-IV record: {'vacant land' if prop_class == '1' else 'low improvement-to-land ratio'}; {acres:.2f} acres; ${land:,} land and ${improvement:,} improvements.",
            "angle": "First verify current ownership, zoning capacity, and environmental constraints; then ask about an option or seller-financed control structure.",
            "status": "Unreviewed",
            "state": "NJ",
            "county": county,
            "sourceName": "NJ Division of Taxation 2026 MOD-IV",
            "sourceUrl": "https://www.nj.gov/treasury/taxation/lpt/statdata.shtml",
            "sourceDate": "2026 assessment list",
            "sourceChecked": CHECKED,
            "sourceNote": "Owner names are redacted in the statewide public export. Building square footage, approvals, tax balance, and environmental constraints still require verification.",
            "_rank": under * .30 + regulatory * .25 + motivation * .20 + strategic * .15 + 50 * .10,
        })
    # Keep a deeper reserve because rural property-location strings do not always
    # resolve through the Census geocoder. The geocoder stops once a county has 12.
    return sorted(rows, key=lambda x: x["_rank"], reverse=True)[:120]


def geocode_nj(rows: list[dict], per_county: int = 12) -> list[dict]:
    cache = json.loads(GEOCODE_CACHE.read_text(encoding="utf-8")) if GEOCODE_CACHE.exists() else {}

    def lookup(key: str) -> tuple[str, str]:
        if key in cache:
            return key, cache[key]
        try:
            params = urllib.parse.urlencode({
                "address": key,
                "benchmark": "Public_AR_Current",
                "format": "json",
            })
            url = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress?" + params
            with urllib.request.urlopen(url, timeout=25) as response:
                matches = json.load(response)["result"]["addressMatches"]
            return key, matches[0]["matchedAddress"].title() if matches else ""
        except Exception:
            return key, ""

    selected: list[dict] = []
    county_counts: dict[str, int] = {}
    counties = sorted({row["county"] for row in rows})
    with ThreadPoolExecutor(max_workers=10) as executor:
        for county in counties:
            county_rows = [row for row in rows if row["county"] == county]
            for start in range(0, len(county_rows), 20):
                batch = county_rows[start : start + 20]
                keyed_rows = [
                    (row, f"{row['addressBase']}, {row['municipality']}, NJ")
                    for row in batch
                ]
                missing = list(dict.fromkeys(key for _row, key in keyed_rows if key not in cache))
                for key, result in executor.map(lookup, missing):
                    cache[key] = result
                for row, key in keyed_rows:
                    standardized = cache.get(key, "")
                    if not re.search(r"\b\d{5}(?:-\d{4})?$", standardized):
                        continue
                    row["address"] = re.sub(r", Nj,", ", NJ,", standardized, flags=re.IGNORECASE)
                    row.pop("addressBase", None)
                    row.pop("municipality", None)
                    selected.append(row)
                    county_counts[county] = county_counts.get(county, 0) + 1
                    if county_counts[county] == per_county:
                        break
                if county_counts.get(county, 0) == per_county:
                    break
    GEOCODE_CACHE.write_text(json.dumps(cache, indent=2), encoding="utf-8")
    short = {county: county_counts.get(county, 0) for county in counties if county_counts.get(county, 0) < 5}
    if short:
        raise RuntimeError(f"NJ counties without five verified addresses: {short}")
    return selected


def arcgis_rows() -> list[dict]:
    fields = (
        "UPI,LOC_ADDRESS,MUNI,OWN1,OWN2,ADDR1,ADDR2,ADDR3,ZIP1,TAXYR,DEED_REC_DATE,"
        "LUC,CLASS,SQFT_PLAN_TOT,ACRE_PLAN_TOT,LOT_ASSESS,PROP_ASSESS,TOT_ASSESS,"
        "FMV319,FMV515,LAST_SALE_PRICE,LEGAL1,LEGAL2,LAND_DEV_ID,SUBDIV_NAME,PLAN_NUM"
    )
    where = (
        "CODE = 1000 AND LOC_ADDRESS IS NOT NULL AND "
        "LUC IN ('V-10','V-11','C-70','C-80','C-96','M-20') AND "
        "TOT_ASSESS >= 25000 AND TOT_ASSESS <= 750000 AND "
        "ACRE_PLAN_TOT >= 0.1 AND ACRE_PLAN_TOT <= 10"
    )
    all_rows: list[dict] = []
    offset = 0
    while True:
        body = urllib.parse.urlencode({
            "where": where,
            "outFields": fields,
            "returnGeometry": "false",
            "resultOffset": offset,
            "resultRecordCount": 2000,
            "orderByFields": "OBJECTID",
            "f": "json",
        }).encode()
        request = urllib.request.Request(CHESTER_PARCELS_URL, data=body)
        with urllib.request.urlopen(request, timeout=60) as response:
            payload = json.load(response)
        features = payload.get("features", [])
        all_rows.extend(feature["attributes"] for feature in features)
        if len(features) < 2000:
            break
        offset += len(features)
    return all_rows


def millis_year(value: int | None) -> int:
    if not value:
        return 0
    try:
        return datetime.fromtimestamp(value / 1000, tz=timezone.utc).year
    except (OSError, OverflowError, ValueError):
        return 0


def chester_candidates() -> list[dict]:
    rows = []
    excluded = (
        "COMMONWEALTH OF", "COUNTY OF CHESTER", "TOWNSHIP", "BOROUGH", "SCHOOL DISTRICT",
        "PECO", "PHILA ELECTRIC", "UTILITIES", "AUTHORITY", "CHURCH", "DIOCESE",
    )
    for raw in arcgis_rows():
        owner = " ".join(x.strip() for x in (raw.get("OWN1"), raw.get("OWN2")) if x and x.strip()) or "Owner not supplied"
        if any(term in owner.upper() for term in excluded):
            continue
        land = number(str(raw.get("LOT_ASSESS") or 0))
        improvement = number(str(raw.get("PROP_ASSESS") or 0))
        total = number(str(raw.get("TOT_ASSESS") or 0))
        lot = number(str(raw.get("SQFT_PLAN_TOT") or 0))
        luc = (raw.get("LUC") or "Not supplied").strip()
        ratio = improvement / land if land else 99
        if luc not in {"V-10", "V-11"} and ratio >= 0.45:
            continue
        municipality = CHESTER_MUNICIPALITIES.get(number(str(raw.get("MUNI") or 0)), "Chester County")
        mailing_parts = [raw.get("ADDR1"), raw.get("ADDR2"), raw.get("ADDR3"), raw.get("ZIP1")]
        mailing = ", ".join(str(x).strip() for x in mailing_parts if x and str(x).strip())
        deed_year_value = millis_year(raw.get("DEED_REC_DATE"))
        plan_ref = (raw.get("PLAN_NUM") or raw.get("LAND_DEV_ID") or "").strip()
        under = 96 if luc in {"V-10", "V-11"} else max(60, min(90, round(92 - ratio * 50)))
        regulatory = 60 if municipality in CHESTER_URBAN_CENTERS else 45
        if plan_ref:
            regulatory += 5
        out_of_state = not bool(re.search(r"\bPA\b", mailing.upper()))
        motivation = 45 + (15 if deed_year_value and deed_year_value <= 2005 else 0) + (15 if out_of_state else 0)
        if owner_type(owner) in {"Business entity", "Estate / trust"}:
            motivation += 5
        strategic = 92 if total <= 100_000 else 87 if total <= 250_000 else 80 if total <= 500_000 else 74
        use_name = LAND_USE_NAMES.get(luc, luc)
        flags = ["Official Chester County parcel", use_name]
        if municipality in CHESTER_URBAN_CENTERS:
            flags.append("County urban center")
        approvals = f"Plan reference {plan_ref} — status not verified" if plan_ref else "No plan reference in parcel layer; not researched"
        age_phrase = f"deed on record dates to {deed_year_value}" if deed_year_value else "deed date not supplied"
        rows.append({
            "id": str(raw.get("UPI") or "").strip(),
            "owner": owner.title(),
            "ownerType": owner_type(owner),
            "addressBase": str(raw.get("LOC_ADDRESS") or "").title(),
            "municipality": municipality,
            "mailing": mailing.title() if mailing else "Not supplied",
            "lot": lot,
            "building": 0,
            "year": 0,
            "land": land,
            "improvement": improvement,
            "zoning": f"Not supplied — assessment use {luc}",
            "flags": flags,
            "delinquency": "Not included in county parcel layer",
            "approvals": approvals,
            "factors": {"underutilization": under, "regulatory": regulatory, "motivation": min(100, motivation), "fit": strategic, "intelligence": 50},
            "reason": f"Official county record: {use_name.lower()}; {lot:,} sf; ${land:,} land and ${improvement:,} improvements; {age_phrase}.",
            "angle": "Confirm municipal zoning and the owner’s timing, then test a low-cost option, seller-financing, or long-close structure.",
            "status": "Unreviewed",
            "state": "PA",
            "county": "Chester",
            "sourceName": "Chester County GIS / Assessment",
            "sourceUrl": "https://www.chesco.org/2198/ChescoViews",
            "sourceDate": f"{raw.get('TAXYR') or 'Current'} assessment record",
            "sourceChecked": CHECKED,
            "sourceNote": "The county parcel layer is updated weekly. Its land-use code is not municipal zoning. Building square footage/year, tax balance, zoning, approval status, and site constraints still require verification.",
            "_rank": under * .30 + regulatory * .25 + motivation * .20 + strategic * .15 + 50 * .10,
        })
    return sorted(rows, key=lambda x: x["_rank"], reverse=True)


def geocode_chester(rows: list[dict]) -> list[dict]:
    cache = json.loads(CHESTER_GEOCODE_CACHE.read_text(encoding="utf-8")) if CHESTER_GEOCODE_CACHE.exists() else {}
    selected = []
    for row in rows[:400]:
        key = f"{row['addressBase']}, {row['municipality']}, PA"
        if key not in cache:
            params = urllib.parse.urlencode({"address": key, "benchmark": "Public_AR_Current", "format": "json"})
            url = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress?" + params
            try:
                with urllib.request.urlopen(url, timeout=20) as response:
                    matches = json.load(response)["result"]["addressMatches"]
                cache[key] = matches[0]["matchedAddress"].title() if matches else ""
            except Exception:
                cache[key] = ""
        standardized = cache[key]
        if not re.search(r"\b\d{5}(?:-\d{4})?$", standardized):
            continue
        row["address"] = standardized
        row.pop("addressBase", None)
        row.pop("municipality", None)
        selected.append(row)
        if len(selected) == 18:
            break
    CHESTER_GEOCODE_CACHE.write_text(json.dumps(cache, indent=2), encoding="utf-8")
    if len(selected) < 18:
        raise RuntimeError(f"Only {len(selected)} Chester County candidates had verified full addresses")
    return selected


def main() -> None:
    previous = json.loads(OUTPUT.read_text(encoding="utf-8")) if OUTPUT.exists() else []
    municipalities = load_nj_municipalities()
    nj_pool = []
    with zipfile.ZipFile(refresh_nj_archive()) as archive:
        county_files = sorted(name for name in archive.namelist() if name.endswith(" 26 RE.txt"))
        if len(county_files) != 21:
            raise RuntimeError(f"Expected 21 New Jersey county files, found {len(county_files)}")
        for name in county_files:
            county = name.removesuffix(" 26 RE.txt")
            with archive.open(name) as raw, io.TextIOWrapper(raw, encoding="latin1") as stream:
                nj_pool.extend(nj_candidates(stream, county, municipalities))
    nj = geocode_nj(nj_pool, per_county=12)
    pa = geocode_chester(chester_candidates())
    records = sorted(nj + pa, key=lambda x: x["_rank"], reverse=True)
    for row in records:
        row.pop("_rank", None)
    # Keep every property that has appeared in an earlier screen. New official
    # candidates are added and refreshed, but an older Keep/Passed/Screened
    # property never vanishes just because it falls outside today's source cut.
    current_ids = {row["id"] for row in records}
    records.extend(row for row in previous if row["id"] not in current_ids)
    # A daily check should not rewrite every property just to change its checked
    # date. Preserve that date when the underlying screened records are identical.
    without_check_date = lambda items: [
        {key: value for key, value in row.items() if key != "sourceChecked"}
        for row in items
    ]
    if previous and without_check_date(previous) == without_check_date(records):
        previous_dates = {row["id"]: row.get("sourceChecked", CHECKED) for row in previous}
        for row in records:
            row["sourceChecked"] = previous_dates.get(row["id"], CHECKED)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(records)} verified public records to {OUTPUT}")


if __name__ == "__main__":
    main()
